import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime

# === Step 1: Define file paths ===
project_folder = Path(r'C:\Users\YOUR_USERNAME\Documents\report_parser_project')
search_file = project_folder / 'report_search_results.xlsx'
template_file = project_folder / 'report_template.xlsx'

# === Step 2: Load the data ===
search_results = pd.read_excel(search_file)
print(f"Loaded search results with {len(search_results)} rows")

# Load the Excel workbook for writing (preserving formatting)
wb = openpyxl.load_workbook(template_file)
ws = wb.active

# Also load as DataFrame for column mapping
template_data = pd.read_excel(template_file)
print(f"Loaded report template with {len(template_data)} rows")

# === Step 3: Define column mapping between the two files ===
column_mapping = {
    'Report Name': 'Business Process / Transaction',
    'Found On Server': 'Matched',
    'Type': 'Report Type',
    'Time taken for report generation': 'Time taken for report generation',
    'Volume of records': 'Volume of records'
}

# Get column letters for headers
headers = {cell.value: cell.column_letter for cell in ws[1] if cell.value}

# Build target column indices
target_columns = {
    search_col: template_data.columns.get_loc(template_col) + 1
    for search_col, template_col in column_mapping.items()
    if template_col in template_data.columns
}

# === Step 4: Perform update ===
print("Transferring data from search results to template...")
updates_made = 0
report_map = {row['Report Name']: i for i, row in search_results.iterrows()}

for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
    report_name = ws.cell(i, target_columns['Report Name']).value

    if report_name in report_map:
        search_row = search_results.iloc[report_map[report_name]]

        for search_col in column_mapping:
            if search_col == 'Report Name' or search_col not in target_columns:
                continue

            cell_value = search_row.get(search_col)
            if pd.notna(cell_value):
                col_letter = get_column_letter(target_columns[search_col])
                ws[f"{col_letter}{i}"] = cell_value
                updates_made += 1

# === Step 5: Save with timestamp ===
timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
output_path = project_folder / f'report_template_updated_{timestamp}.xlsx'
wb.save(output_path)

print(f"Update complete! Made {updates_made} cell updates.")
print(f"Updated file saved to: {output_path}")

# === Step 6: Optional: Show sample ===
print("\nSample of updated values:")
df_updated = pd.read_excel(output_path)
sample = df_updated.sample(min(5, len(df_updated)))
print(sample[['Business Process / Transaction', 'Report Type', 'Time taken for report generation', 'Volume of records']])
